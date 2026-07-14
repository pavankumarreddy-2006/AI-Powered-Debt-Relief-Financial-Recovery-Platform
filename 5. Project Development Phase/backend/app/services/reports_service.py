import io
from datetime import datetime
from typing import List, Dict, Any
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_pdf_report(user: Any, loans: List[Any], recommendations: List[Any], analysis: Any) -> bytes:
    """Generates a professional financial recovery PDF report in-memory."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    primary_color = colors.HexColor("#2563EB")   # Slate Blue
    secondary_color = colors.HexColor("#10B981") # Emerald Green
    neutral_dark = colors.HexColor("#1F2937")    # Dark Charcoal
    neutral_light = colors.HexColor("#F3F4F6")   # Soft Grey
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=primary_color,
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=25
    )
    
    h1_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=primary_color,
        spaceBefore=15,
        spaceAfter=10,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10.5,
        textColor=neutral_dark,
        leading=14,
        spaceAfter=8
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1 # Center
    )
    
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        textColor=neutral_dark,
        alignment=0 # Left
    )

    table_cell_center = ParagraphStyle(
        'TableCellCenter',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        textColor=neutral_dark,
        alignment=1 # Center
    )

    story = []
    
    # Header Section
    story.append(Paragraph("FINANCIAL RECOVERY & DEBT RELIEF REPORT", title_style))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y')} | Platform: AI Powered Debt Relief", subtitle_style))
    
    # User Profile Section
    story.append(Paragraph("1. Client Profile & Overview", h1_style))
    profile_data = [
        [Paragraph("<b>Full Name:</b>", body_style), Paragraph(user.full_name, body_style), Paragraph("<b>Email:</b>", body_style), Paragraph(user.email, body_style)],
        [Paragraph("<b>Occupation:</b>", body_style), Paragraph(user.occupation or "Not Provided", body_style), Paragraph("<b>Phone:</b>", body_style), Paragraph(user.phone or "Not Provided", body_style)],
        [Paragraph("<b>Monthly Income:</b>", body_style), Paragraph(f"${user.monthly_income:,.2f}", body_style), Paragraph("<b>Analysis Date:</b>", body_style), Paragraph(datetime.now().strftime("%Y-%m-%d"), body_style)]
    ]
    t_profile = Table(profile_data, colWidths=[1.2*inch, 2.2*inch, 1.2*inch, 2.2*inch])
    t_profile.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('BACKGROUND', (0,0), (-1,-1), neutral_light),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#E5E7EB")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ('TOPPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t_profile)
    story.append(Spacer(1, 15))
    
    # Financial Health Metrics
    if analysis:
        story.append(Paragraph("2. Financial Health Indicators", h1_style))
        indicator_color = colors.HexColor("#EF4444")  # Red
        if analysis.status_indicator == "Green":
            indicator_color = colors.HexColor("#10B981")
        elif analysis.status_indicator == "Yellow":
            indicator_color = colors.HexColor("#F59E0B")
            
        health_data = [
            [
                Paragraph("<b>Debt-to-Income (DTI) Ratio:</b>", body_style), Paragraph(f"{analysis.debt_to_income_ratio * 100:.1f}%", body_style),
                Paragraph("<b>Monthly Surplus:</b>", body_style), Paragraph(f"${analysis.monthly_surplus:,.2f}", body_style)
            ],
            [
                Paragraph("<b>Stability Score:</b>", body_style), Paragraph(f"{analysis.stability_score:.1f} / 100", body_style),
                Paragraph("<b>Debt Stress Score:</b>", body_style), Paragraph(f"{analysis.debt_stress_index:.1f} / 100", body_style)
            ],
            [
                Paragraph("<b>Hardship Status:</b>", body_style), Paragraph(f"<font color='{indicator_color}'><b>{analysis.status_indicator.upper()}</b></font>", body_style),
                Paragraph("<b>Disposable Income:</b>", body_style), Paragraph(f"${analysis.disposable_income:,.2f}", body_style)
            ]
        ]
        t_health = Table(health_data, colWidths=[2.0*inch, 1.4*inch, 1.6*inch, 1.8*inch])
        t_health.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BACKGROUND', (0,0), (-1,-1), colors.white),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#D1D5DB")),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ]))
        story.append(t_health)
        story.append(Spacer(1, 15))
        
    # Outstanding Loans Table
    story.append(Paragraph("3. Outstanding Liability Ledger", h1_style))
    if not loans:
        story.append(Paragraph("No loans registered on this profile.", body_style))
    else:
        loan_headers = ["Loan Name", "Lender", "Type", "Interest", "Original", "Outstanding", "EMI", "Status"]
        loan_table_data = [[Paragraph(f"<b>{h}</b>", table_header_style) for h in loan_headers]]
        
        for l in loans:
            loan_table_data.append([
                Paragraph(l.loan_name, table_cell_style),
                Paragraph(l.lender_name, table_cell_style),
                Paragraph(l.loan_type, table_cell_style),
                Paragraph(f"{l.interest_rate}%", table_cell_center),
                Paragraph(f"${l.original_amount:,.2f}", table_cell_center),
                Paragraph(f"${l.outstanding_amount:,.2f}", table_cell_center),
                Paragraph(f"${l.emi:,.2f}", table_cell_center),
                Paragraph(l.status, table_cell_center)
            ])
            
        t_loans = Table(loan_table_data, colWidths=[1.1*inch, 1.1*inch, 0.8*inch, 0.6*inch, 0.9*inch, 0.9*inch, 0.7*inch, 0.7*inch])
        t_loans.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), primary_color),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D1D5DB")),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, neutral_light])
        ]))
        story.append(t_loans)
    story.append(Spacer(1, 20))
    
    # AI Recommendations Section (Keep together or page break if necessary)
    if recommendations:
        story.append(Paragraph("4. AI-Generated Settlement Recommendations", h1_style))
        rec_headers = ["Target Loan", "Lender", "Suggested Sett. %", "Suggested Payment", "Probability", "Risk Level"]
        rec_table_data = [[Paragraph(f"<b>{h}</b>", table_header_style) for h in rec_headers]]
        
        # Build dictionary map of loans for quick lookups
        loan_map = {l.id: l for l in loans}
        
        for r in recommendations:
            loan_obj = loan_map.get(r.loan_id)
            loan_name = loan_obj.loan_name if loan_obj else "Unknown"
            lender_name = loan_obj.lender_name if loan_obj else "Unknown"
            
            rec_table_data.append([
                Paragraph(loan_name, table_cell_style),
                Paragraph(lender_name, table_cell_style),
                Paragraph(f"{r.recommended_settlement_pct}% ({r.expected_range})", table_cell_center),
                Paragraph(f"${r.recommended_monthly_payment:,.2f}", table_cell_center),
                Paragraph(f"{r.settlement_probability * 100:.0f}%", table_cell_center),
                Paragraph(r.risk_level, table_cell_center)
            ])
            
        t_recs = Table(rec_table_data, colWidths=[1.3*inch, 1.3*inch, 1.4*inch, 1.2*inch, 0.8*inch, 0.8*inch])
        t_recs.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), secondary_color),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D1D5DB")),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, neutral_light])
        ]))
        story.append(t_recs)
        
        # Include detailed recovery advice from the latest recommendation
        story.append(Spacer(1, 10))
        latest_rec = recommendations[-1]
        story.append(Paragraph("<b>AI Strategic Action Plan:</b>", body_style))
        story.append(Paragraph(latest_rec.recovery_advice, body_style))
        
    story.append(Spacer(1, 20))
    story.append(Paragraph("5. Hardship and Legal Disclaimers", h1_style))
    disclaimer_text = (
        "Disclaimer: This document contains analytical tools, financial calculations, and AI-generated "
        "recommendations intended for informational and educational purposes. The contents do not constitute "
        "formal legal or tax advice. Debt settlement strategies carry risks including credit score impacts, "
        "potential collection efforts, or tax liabilities for forgiven debts. Users are urged to consult with "
        "licensed debt advisors, bankruptcy attorneys, or legal experts before executing critical settlement actions."
    )
    story.append(Paragraph(disclaimer_text, ParagraphStyle('Disclaimer', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=8, textColor=colors.HexColor("#7F8C8D"), leading=11)))
    
    doc.build(story)
    return buffer.getvalue()

def generate_excel_report(user: Any, loans: List[Any], recommendations: List[Any], analysis: Any) -> bytes:
    """Generates a professional financial recovery Excel spreadsheet in-memory."""
    wb = openpyxl.Workbook()
    
    # Formatting styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="2563EB")
    section_font = Font(name=font_family, size=12, bold=True, color="1F2937")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    bold_cell_font = Font(name=font_family, size=10, bold=True)
    normal_font = Font(name=font_family, size=10)
    italic_font = Font(name=font_family, size=9, italic=True, color="7F8C8D")
    
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_fill_emerald = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    grey_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    double_bottom_border = Border(
        bottom=Side(style='double', color='1F2937'),
        top=Side(style='thin', color='D1D5DB')
    )

    # ================= SHEET 1: OVERVIEW =================
    ws_ov = wb.active
    ws_ov.title = "Financial Overview"
    ws_ov.views.sheetView[0].showGridLines = True
    
    ws_ov.cell(row=2, column=2, value="FINANCIAL HEALTH & DEBT RECOVERY OVERVIEW").font = title_font
    ws_ov.cell(row=3, column=2, value=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = italic_font
    
    # Client section
    ws_ov.cell(row=5, column=2, value="Borrower Details").font = section_font
    ws_ov.cell(row=6, column=2, value="Full Name:").font = bold_cell_font
    ws_ov.cell(row=6, column=3, value=user.full_name).font = normal_font
    ws_ov.cell(row=7, column=2, value="Email:").font = bold_cell_font
    ws_ov.cell(row=7, column=3, value=user.email).font = normal_font
    ws_ov.cell(row=8, column=2, value="Occupation:").font = bold_cell_font
    ws_ov.cell(row=8, column=3, value=user.occupation or "Not Provided").font = normal_font
    
    # Analysis Metrics
    if analysis:
        ws_ov.cell(row=5, column=5, value="Key Financial Health Ratios").font = section_font
        
        metrics = [
            ("Debt-to-Income (DTI)", analysis.debt_to_income_ratio, "0.0%"),
            ("Monthly Income Surplus", analysis.monthly_surplus, "$#,##0.00"),
            ("Remaining Disposable Income", analysis.disposable_income, "$#,##0.00"),
            ("Debt Stress index", analysis.debt_stress_index / 100, "0.0%"),
            ("Financial Stability Score", analysis.stability_score / 100, "0.0%"),
            ("Hardship Category Indicator", analysis.status_indicator, "@")
        ]
        
        for idx, (label, val, num_fmt) in enumerate(metrics, start=6):
            ws_ov.cell(row=idx, column=5, value=label).font = bold_cell_font
            cell_val = ws_ov.cell(row=idx, column=6, value=val)
            cell_val.font = normal_font
            cell_val.number_format = num_fmt
            
    # Add borders to borrower and metric cards
    for r in range(6, 9):
        ws_ov.cell(row=r, column=2).border = thin_border
        ws_ov.cell(row=r, column=3).border = thin_border
    if analysis:
        for r in range(6, 12):
            ws_ov.cell(row=r, column=5).border = thin_border
            ws_ov.cell(row=r, column=6).border = thin_border

    # ================= SHEET 2: LOANS =================
    ws_ln = wb.create_sheet(title="Loan Ledger")
    ws_ln.views.sheetView[0].showGridLines = True
    
    ws_ln.cell(row=2, column=2, value="Liability & Loan Registry").font = section_font
    
    headers = ["Loan ID", "Loan Name", "Lender Name", "Loan Type", "Original Amount", "Outstanding Amount", "Interest Rate (%)", "Monthly EMI", "Due Date", "Overdue Months", "Status"]
    
    for c_idx, h in enumerate(headers, start=2):
        cell = ws_ln.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    # Populating rows
    tot_orig = 0
    tot_out = 0
    tot_emi = 0
    
    for r_idx, l in enumerate(loans, start=5):
        ws_ln.cell(row=r_idx, column=2, value=l.id).font = normal_font
        ws_ln.cell(row=r_idx, column=3, value=l.loan_name).font = normal_font
        ws_ln.cell(row=r_idx, column=4, value=l.lender_name).font = normal_font
        ws_ln.cell(row=r_idx, column=5, value=l.loan_type).font = normal_font
        
        orig = ws_ln.cell(row=r_idx, column=6, value=l.original_amount)
        orig.font = normal_font
        orig.number_format = "$#,##0.00"
        tot_orig += l.original_amount
        
        out = ws_ln.cell(row=r_idx, column=7, value=l.outstanding_amount)
        out.font = normal_font
        out.number_format = "$#,##0.00"
        tot_out += l.outstanding_amount
        
        rate = ws_ln.cell(row=r_idx, column=8, value=l.interest_rate / 100)
        rate.font = normal_font
        rate.number_format = "0.0%"
        
        emi = ws_ln.cell(row=r_idx, column=9, value=l.emi)
        emi.font = normal_font
        emi.number_format = "$#,##0.00"
        tot_emi += l.emi
        
        ws_ln.cell(row=r_idx, column=10, value=l.due_date).font = normal_font
        ws_ln.cell(row=r_idx, column=11, value=l.overdue_months).font = normal_font
        ws_ln.cell(row=r_idx, column=12, value=l.status).font = normal_font
        
        # Border cells
        for c in range(2, 13):
            ws_ln.cell(row=r_idx, column=c).border = thin_border
            
    # Totals Row
    total_row = len(loans) + 5
    ws_ln.cell(row=total_row, column=5, value="Total Summary:").font = bold_cell_font
    
    orig_total = ws_ln.cell(row=total_row, column=6, value=tot_orig)
    orig_total.font = bold_cell_font
    orig_total.number_format = "$#,##0.00"
    
    out_total = ws_ln.cell(row=total_row, column=7, value=tot_out)
    out_total.font = bold_cell_font
    out_total.number_format = "$#,##0.00"
    
    emi_total = ws_ln.cell(row=total_row, column=9, value=tot_emi)
    emi_total.font = bold_cell_font
    emi_total.number_format = "$#,##0.00"
    
    for c in range(2, 13):
        ws_ln.cell(row=total_row, column=c).border = double_bottom_border

    # ================= SHEET 3: SETTLEMENTS =================
    ws_st = wb.create_sheet(title="Settlement Strategies")
    ws_st.views.sheetView[0].showGridLines = True
    
    ws_st.cell(row=2, column=2, value="AI Recommended Settlement Matrices").font = section_font
    
    rec_headers = ["Loan ID", "Loan Name", "Lender", "Outstanding Balance", "Recommended Settlement (%)", "Settlement Amount Offer", "Suggested Monthly Payment", "Settlement Likelihood", "Stress level", "Risk Category"]
    
    for c_idx, h in enumerate(rec_headers, start=2):
        cell = ws_st.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill_emerald
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    loan_map = {l.id: l for l in loans}
    
    for r_idx, r in enumerate(recommendations, start=5):
        loan_obj = loan_map.get(r.loan_id)
        l_name = loan_obj.loan_name if loan_obj else "Unknown"
        l_lender = loan_obj.lender_name if loan_obj else "Unknown"
        l_bal = loan_obj.outstanding_amount if loan_obj else 0.0
        
        ws_st.cell(row=r_idx, column=2, value=r.loan_id).font = normal_font
        ws_st.cell(row=r_idx, column=3, value=l_name).font = normal_font
        ws_st.cell(row=r_idx, column=4, value=l_lender).font = normal_font
        
        bal_cell = ws_st.cell(row=r_idx, column=5, value=l_bal)
        bal_cell.font = normal_font
        bal_cell.number_format = "$#,##0.00"
        
        pct_cell = ws_st.cell(row=r_idx, column=6, value=r.recommended_settlement_pct / 100)
        pct_cell.font = normal_font
        pct_cell.number_format = "0.0%"
        
        amt_cell = ws_st.cell(row=r_idx, column=7, value=l_bal * (r.recommended_settlement_pct / 100))
        amt_cell.font = normal_font
        amt_cell.number_format = "$#,##0.00"
        
        pmt_cell = ws_st.cell(row=r_idx, column=8, value=r.recommended_monthly_payment)
        pmt_cell.font = normal_font
        pmt_cell.number_format = "$#,##0.00"
        
        prob_cell = ws_st.cell(row=r_idx, column=9, value=r.settlement_probability)
        prob_cell.font = normal_font
        prob_cell.number_format = "0.0%"
        
        ws_st.cell(row=r_idx, column=10, value=r.stress_level).font = normal_font
        ws_st.cell(row=r_idx, column=11, value=r.risk_level).font = normal_font
        
        for c in range(2, 12):
            ws_st.cell(row=r_idx, column=c).border = thin_border
            
    # Auto-adjust column widths for all sheets
    for ws in [ws_ov, ws_ln, ws_st]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    return out_buffer.getvalue()
